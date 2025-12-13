# applications/finanzas/views.py

"""
Módulo de Vistas para la aplicación 'finanzas'.

Este archivo contiene la lógica de presentación y procesamiento de datos para el módulo
de gestión financiera. Incluye vistas para la renderización del dashboard principal,
endpoints de API para la alimentación de gráficos (Chart.js), procesamiento de formularios
de gastos y la generación de reportes exportables en formato Excel.
"""

import json
from datetime import datetime, date, time
from django.utils import timezone
from django.views.generic import View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse # <-- ¡NUEVO!
from django.db.models import Sum, F, Count, DecimalField, Case, When, Value
from django.db.models.functions import TruncMonth, TruncDay, ExtractHour
from django.conf import settings

# --- ¡NUEVO! Importaciones para Excel ---
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
# -------------------------------------

from applications.ventas.models import Venta, DetalleVenta
from applications.stock.models import Categoria as CategoriaProducto, Producto
from .models import Gasto, CategoriaGasto
from .forms import GastoForm, CategoriaGastoForm

# VISTA 1: La que renderiza el HTML
# ------------------------------------------------
class DashboardFinanzasView(LoginRequiredMixin, TemplateView):
    """
    Vista basada en plantilla que renderiza la interfaz principal del Dashboard Financiero.

    Se encarga de inicializar el contexto con los formularios necesarios para el registro
    de gastos y categorías, así como establecer los rangos de fechas predeterminados
    para la visualización inicial de los datos.
    """
    template_name = 'finanzas/dashboard.html'
    
    def get_context_data(self, **kwargs):
        """
        Prepara el contexto de datos para la plantilla.
        
        Inicializa las fechas de filtro (inicio de mes actual hasta hoy) e instancia
        los formularios de registro de gastos y categorías para su uso en modales.
        """
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Dashboard Financiero'
        today = timezone.now().date()
        context['fecha_hoy'] = today
        context['fecha_inicio_mes'] = today.replace(day=1)
        context['gasto_form'] = GastoForm()
        context['categoria_gasto_form'] = CategoriaGastoForm()
        return context


# VISTA 2: La que provee los datos (JSON)
# ------------------------------------------------
class FinanzasDataJSONView(LoginRequiredMixin, View):
    """
    Vista de API que provee datos financieros en formato JSON.

    Esta vista es consumida asíncronamente por el frontend para renderizar los gráficos
    y actualizar los indicadores clave de rendimiento (KPIs) sin recargar la página.
    Realiza cálculos agregados sobre ventas, costos y gastos en un rango de fechas específico.
    """
    
    def get(self, request, *args, **kwargs):
        """
        Maneja las solicitudes GET para la obtención de datos financieros.

        Procesa los parámetros de filtrado por fecha, calcula métricas de rentabilidad
        (ingresos, COGS, beneficio neto) y estructura los datos para su visualización
        en gráficos de líneas, barras y torta.
        """
        
        # --- 1. Obtener y validar el rango de fechas ---
        try:
            today = timezone.now().date()
            start_date_str = request.GET.get('start', today.replace(day=1).strftime('%Y-%m-%d'))
            end_date_str = request.GET.get('end', today.strftime('%Y-%m-%d'))
            
            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            # Fechas para consultas de Gastos (DateField)
            start_date_for_gasto = start_date_obj
            end_date_for_gasto = end_date_obj
            
            # Fechas para consultas de Ventas (DateTimeField - con manejo de zona horaria)
            start_dt_naive = datetime.combine(start_date_obj, time.min)
            end_dt_naive = datetime.combine(end_date_obj, time.max)

            start_date_aware = timezone.make_aware(start_dt_naive, timezone.get_current_timezone())
            end_date_aware = timezone.make_aware(end_dt_naive, timezone.get_current_timezone())

        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido. Usar YYYY-MM-DD.'}, status=400)

        # --- 2. Calcular los KPIs Principales ---
        
        # Cálculo de Ingresos Brutos (Suma total de ventas en el período)
        total_ingresos = Venta.objects.filter(
            fecha_hora__range=[start_date_aware, end_date_aware]
        ).aggregate(total=Sum('total'))['total'] or 0
        
        # Cálculo del Costo de Mercadería Vendida (COGS) basado en el costo histórico al momento de la venta
        total_cogs = DetalleVenta.objects.filter(
            venta__fecha_hora__range=[start_date_aware, end_date_aware]
        ).aggregate(total=Sum(F('cantidad') * F('precio_compra_momento')))['total'] or 0
        
        # Cálculo de Gastos Operativos imputados en el período
        total_gastos = Gasto.objects.filter(
            fecha_imputacion__range=[start_date_for_gasto, end_date_for_gasto]
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        # Cálculo de Resultados
        ganancia_bruta = total_ingresos - total_cogs
        beneficio_neto = ganancia_bruta - total_gastos

        # --- 3. Preparar datos para los Gráficos (Torta y Líneas) ---
        
        # Datos para Gráfico de Gastos por Categoría
        gastos_por_categoria_qs = CategoriaGasto.objects.filter(
            gasto__fecha_imputacion__range=[start_date_for_gasto, end_date_for_gasto]
        ).annotate(total=Sum('gasto__monto')).order_by('-total')
        chart_gastos_categoria = {
            'labels': [g.nombre for g in gastos_por_categoria_qs if g.total > 0],
            'data': [g.total for g in gastos_por_categoria_qs if g.total > 0],
        }

        # Datos para Gráfico de Ventas por Categoría de Producto
        ventas_por_categoria_qs = CategoriaProducto.objects.filter(
            productos__detalleventa__venta__fecha_hora__range=[start_date_aware, end_date_aware]
        ).annotate(total_vendido=Sum('productos__detalleventa__subtotal')).order_by('-total_vendido')
        chart_ventas_categoria = {
            'labels': [c.nombre for c in ventas_por_categoria_qs if c.total_vendido > 0],
            'data': [c.total_vendido for c in ventas_por_categoria_qs if c.total_vendido > 0],
        }

        # Datos para Gráfico de Evolución Temporal (Ingresos vs Gastos diarios)
        ingresos_por_dia = Venta.objects.filter(
            fecha_hora__range=[start_date_aware, end_date_aware]
        ).annotate(dia=TruncDay('fecha_hora')).values('dia').annotate(total=Sum('total')).order_by('dia')
        
        gastos_por_dia = Gasto.objects.filter(
            fecha_imputacion__range=[start_date_for_gasto, end_date_for_gasto]
        ).annotate(dia=TruncDay('fecha_imputacion')).values('dia').annotate(total=Sum('monto')).order_by('dia')
        
        # Normalización de series temporales para alinear fechas en el eje X
        ingresos_map = {d['dia'].strftime('%Y-%m-%d'): d['total'] for d in ingresos_por_dia}
        gastos_map = {d['dia'].strftime('%Y-%m-%d'): d['total'] for d in gastos_por_dia}
        all_labels = sorted(list(set(ingresos_map.keys()) | set(gastos_map.keys())))
        
        chart_evolucion = {
            'labels': all_labels,
            'ingresos_data': [ingresos_map.get(label, 0) for label in all_labels],
            'gastos_data': [gastos_map.get(label, 0) for label in all_labels],
        }

        # --- 4. Cálculos para Rankings (Top productos y vendedores) ---
        
        detalles_en_rango = DetalleVenta.objects.filter(
            venta__fecha_hora__range=[start_date_aware, end_date_aware]
        ).select_related('producto')
        
        # Top 5 Productos más vendidos por cantidad
        top_productos_venta_qs = detalles_en_rango.values(
            'producto__nombre'
        ).annotate(
            total_cantidad=Sum('cantidad')
        ).order_by('-total_cantidad')[:5]
        chart_top_productos_venta = {
            'labels': [p['producto__nombre'] for p in top_productos_venta_qs],
            'data': [p['total_cantidad'] for p in top_productos_venta_qs],
        }
        
        # Top 5 Productos más rentables (Margen de ganancia)
        top_productos_rentables_qs = detalles_en_rango.annotate(
            ganancia_linea=F('subtotal') - (F('cantidad') * F('precio_compra_momento'))
        ).values('producto__nombre').annotate(
            ganancia_total=Sum('ganancia_linea')
        ).order_by('-ganancia_total')[:5]
        chart_top_productos_rentables = {
            'labels': [p['producto__nombre'] for p in top_productos_rentables_qs],
            'data': [p['ganancia_total'] for p in top_productos_rentables_qs],
        }
        
        # Ventas acumuladas por Vendedor
        ventas_por_vendedor_qs = Venta.objects.filter(
            fecha_hora__range=[start_date_aware, end_date_aware]
        ).values('vendedor__username').annotate(
            total_vendido=Sum('total')
        ).order_by('-total_vendido')
        chart_ventas_vendedor = {
            'labels': [v['vendedor__username'] or 'Sin Vendedor' for v in ventas_por_vendedor_qs],
            'data': [v['total_vendido'] for v in ventas_por_vendedor_qs],
        }
        
        # --- 5. Análisis de Horarios y Turnos ---
        
        base_ventas_rango = Venta.objects.filter(
            fecha_hora__range=[start_date_aware, end_date_aware]
        )
        
        # Distribución de ventas por hora del día
        ventas_por_hora_qs = base_ventas_rango.annotate(
            hora=ExtractHour('fecha_hora')
        ).values('hora').annotate(
            total=Sum('total')
        ).order_by('hora')
        ventas_por_hora_map = {f"{h:02d}:00": 0 for h in range(24)}
        for v in ventas_por_hora_qs:
            ventas_por_hora_map[f"{v['hora']:02d}:00"] = v['total']
        chart_ventas_hora = {
            'labels': list(ventas_por_hora_map.keys()),
            'data': list(ventas_por_hora_map.values()),
        }
        
        # Distribución de ventas por turno (Mañana, Tarde, Noche)
        ventas_por_turno_qs = base_ventas_rango.annotate(
            hora=ExtractHour('fecha_hora')
        ).annotate(
            turno=Case(
                When(hora__gte=7, hora__lt=15, then=Value('Mañana (7-15hs)')),
                When(hora__gte=15, hora__lt=24, then=Value('Tarde (15-00hs)')),
                default=Value('Noche (00-7hs)')
            )
        ).values('turno').annotate(
            total=Sum('total')
        ).order_by('turno')
        chart_ventas_turno = {
            'labels': [t['turno'] for t in ventas_por_turno_qs],
            'data': [t['total'] for t in ventas_por_turno_qs],
        }

        # --- 6. Construir la respuesta JSON ---
        data = {
            'kpis': { 'total_ingresos': total_ingresos, 'total_cogs': total_cogs, 'ganancia_bruta': ganancia_bruta, 'total_gastos': total_gastos, 'beneficio_neto': beneficio_neto, },
            'charts': {
                'gastos_por_categoria': chart_gastos_categoria, 'ventas_por_categoria': chart_ventas_categoria, 'evolucion_ingresos_gastos': chart_evolucion,
                'top_productos_venta': chart_top_productos_venta, 'top_productos_rentables': chart_top_productos_rentables, 'ventas_por_vendedor': chart_ventas_vendedor,
                'ventas_por_hora': chart_ventas_hora, 'ventas_por_turno': chart_ventas_turno,
            }
        }
        return JsonResponse(data)


# VISTA 3: La que guarda el gasto
# ------------------------------------------------
class RegistrarGastoView(LoginRequiredMixin, View):
    """
    Vista funcional para el registro de nuevos gastos operativos.

    Procesa solicitudes POST enviadas desde el formulario modal del dashboard.
    Asigna automáticamente el usuario que registra la operación para mantener
    la trazabilidad.
    """
    def post(self, request, *args, **kwargs):
        form = GastoForm(request.POST)
        if form.is_valid():
            gasto = form.save(commit=False)
            gasto.usuario_registra = request.user
            gasto.save()
            return JsonResponse({'success': True, 'gasto_id': gasto.id})
        else:
            return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)

# VISTA 4: La que guarda la categoría
# ------------------------------------------------
class RegistrarCategoriaGastoView(LoginRequiredMixin, View):
    """
    Vista funcional para la creación dinámica de categorías de gastos.

    Permite añadir nuevas categorías directamente desde la interfaz de registro de gastos,
    mejorando la usabilidad al evitar interrupciones en el flujo de trabajo.
    """
    def post(self, request, *args, **kwargs):
        form = CategoriaGastoForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            return JsonResponse({ 'success': True, 'categoria_id': categoria.id, 'categoria_nombre': categoria.nombre, })
        else:
            return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)


# VISTA 5: VISTA PARA EXPORTAR EXCEL
# ------------------------------------------------
class ExportarExcelView(LoginRequiredMixin, View):
    """
    Vista encargada de la generación y descarga de reportes financieros en formato Excel (.xlsx).

    Utiliza la librería `openpyxl` para construir un archivo de hoja de cálculo en memoria
    que contiene:
    1. Una hoja de resumen con los KPIs principales del período seleccionado.
    2. Una hoja con el detalle transaccional de todas las ventas.
    3. Una hoja con el detalle de todos los gastos operativos registrados.
    """
    
    def get(self, request, *args, **kwargs):
        """
        Procesa la solicitud de descarga del reporte.

        Recupera el rango de fechas, consulta la base de datos para obtener ventas y gastos,
        y formatea la información en hojas de Excel con estilos y encabezados apropiados.
        Retorna el archivo binario como una respuesta HTTP descargable.
        """
        
        # --- 1. Obtener y validar el rango de fechas ---
        try:
            today = timezone.now().date()
            start_date_str = request.GET.get('start', today.replace(day=1).strftime('%Y-%m-%d'))
            end_date_str = request.GET.get('end', today.strftime('%Y-%m-%d'))
            
            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            start_date_for_gasto = start_date_obj
            end_date_for_gasto = end_date_obj
            
            start_dt_naive = datetime.combine(start_date_obj, time.min)
            end_dt_naive = datetime.combine(end_date_obj, time.max)

            start_date_aware = timezone.make_aware(start_dt_naive, timezone.get_current_timezone())
            end_date_aware = timezone.make_aware(end_dt_naive, timezone.get_current_timezone())
        except ValueError:
            return HttpResponse("Error en formato de fecha. Use YYYY-MM-DD.", status=400)

        # --- 2. Crear el libro de Excel en memoria ---
        wb = openpyxl.Workbook()
        
        # Estilos predefinidos para encabezados
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center')
        
        # --- 3. Pestaña 1: Resumen de KPIs ---
        ws_resumen = wb.active
        ws_resumen.title = "Resumen (KPIs)"
        
        # Consultamos los KPIs (Misma lógica que en la vista de JSON)
        total_ingresos = Venta.objects.filter(fecha_hora__range=[start_date_aware, end_date_aware]).aggregate(total=Sum('total'))['total'] or 0
        total_cogs = DetalleVenta.objects.filter(venta__fecha_hora__range=[start_date_aware, end_date_aware]).aggregate(total=Sum(F('cantidad') * F('precio_compra_momento')))['total'] or 0
        total_gastos = Gasto.objects.filter(fecha_imputacion__range=[start_date_for_gasto, end_date_for_gasto]).aggregate(total=Sum('monto'))['total'] or 0
        ganancia_bruta = total_ingresos - total_cogs
        beneficio_neto = ganancia_bruta - total_gastos

        # Escritura de datos en la hoja de resumen
        ws_resumen.append(['Reporte Financiero'])
        ws_resumen.append([f"Desde: {start_date_obj.strftime('%d/%m/%Y')} hasta {end_date_obj.strftime('%d/%m/%Y')}"])
        ws_resumen.append([]) # Fila vacía
        ws_resumen.append(['Métrica', 'Valor'])
        ws_resumen.append(['Ingresos Brutos', total_ingresos])
        ws_resumen.append(['Costo de Mercadería (COGS)', total_cogs])
        ws_resumen.append(['Ganancia Bruta', ganancia_bruta])
        ws_resumen.append(['Gastos Operativos', total_gastos])
        ws_resumen.append(['BENEFICIO NETO', beneficio_neto])

        # Aplicar estilos al resumen
        ws_resumen['A1'].font = Font(bold=True, size=16)
        for cell in ws_resumen['A4:B4']: cell[0].font = header_font
        for cell in ws_resumen['A9:B9']: cell[0].font = header_font
        for col in ['A', 'B']: ws_resumen.column_dimensions[col].width = 30
        
        # --- 4. Pestaña 2: Detalle de Ventas (Datos Crudos) ---
        ws_ventas = wb.create_sheet(title="Ventas (Detallado)")
        
        # Encabezados de la tabla de ventas
        headers_ventas = [
            'ID Venta', 'Fecha/Hora', 'Vendedor', 'Producto', 'Categoría', 
            'Cantidad', 'Precio Unit.', 'Subtotal (Venta)', 'Costo Unit.', 'Costo Total', 'Ganancia'
        ]
        ws_ventas.append(headers_ventas)
        
        # Aplicar estilos a encabezados
        for col_num, header in enumerate(headers_ventas, 1):
            cell = ws_ventas.cell(row=1, column=col_num)
            cell.font = header_font
            ws_ventas.column_dimensions[get_column_letter(col_num)].width = 20

        # Consultar y poblar datos detallados de ventas
        detalles_qs = DetalleVenta.objects.filter(
            venta__fecha_hora__range=[start_date_aware, end_date_aware]
        ).select_related(
            'venta', 'producto', 'venta__vendedor', 'producto__categoria'
        ).order_by('venta__fecha_hora')

        for detalle in detalles_qs:
            costo_total_linea = detalle.cantidad * detalle.precio_compra_momento
            ganancia_linea = detalle.subtotal - costo_total_linea
            
            # Formatear fecha para Excel (sin zona horaria para compatibilidad)
            fecha_local = timezone.localtime(detalle.venta.fecha_hora)
            
            ws_ventas.append([
                detalle.venta.id,
                fecha_local.strftime("%Y-%m-%d %H:%M:%S"),
                detalle.venta.vendedor.username if detalle.venta.vendedor else 'N/A',
                detalle.producto.nombre if detalle.producto else 'N/A',
                detalle.producto.categoria.nombre if detalle.producto and detalle.producto.categoria else 'N/A',
                detalle.cantidad,
                detalle.precio_unitario_momento,
                detalle.subtotal,
                detalle.precio_compra_momento,
                costo_total_linea,
                ganancia_linea
            ])

        # --- 5. Pestaña 3: Detalle de Gastos (Datos Crudos) ---
        ws_gastos = wb.create_sheet(title="Gastos (Detallado)")
        
        headers_gastos = [
            'ID Gasto', 'Fecha Imputación', 'Categoría', 'Monto', 'Descripción', 'Registrado Por'
        ]
        ws_gastos.append(headers_gastos)
        
        for col_num, header in enumerate(headers_gastos, 1):
            cell = ws_gastos.cell(row=1, column=col_num)
            cell.font = header_font
            ws_gastos.column_dimensions[get_column_letter(col_num)].width = 25
            
        gastos_qs = Gasto.objects.filter(
            fecha_imputacion__range=[start_date_for_gasto, end_date_for_gasto]
        ).select_related('categoria', 'usuario_registra').order_by('fecha_imputacion')
        
        for gasto in gastos_qs:
            ws_gastos.append([
                gasto.id,
                gasto.fecha_imputacion.strftime("%Y-%m-%d"),
                gasto.categoria.nombre if gasto.categoria else 'N/A',
                gasto.monto,
                gasto.descripcion,
                gasto.usuario_registra.username if gasto.usuario_registra else 'N/A'
            ])

        # --- 6. Guardar el libro en un buffer y crear la respuesta ---
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Reporte_Finanzas_{start_date_obj.strftime('%Y-%m-%d')}_al_{end_date_obj.strftime('%Y-%m-%d')}.xlsx"
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response